"""
Generate publications/publications.json from publications/my_publications.xlsx.

Usage:
    python generate_publications.py

The Excel file is the single source of truth for publication data.
Both HTML pages (publications.html, research.html) load the generated JSON via fetch().
"""

import json
import re
import openpyxl

EXCEL_PATH = "../publications/my_publications.xlsx"
JSON_PATH = "../publications/publications.json"


def normalize_text(s):
    """Strip and replace non-breaking spaces with regular spaces."""
    if not s:
        return ""
    return s.replace("\xa0", " ").strip()


def normalize_authors(s):
    """Normalize author string to consistent 'Last I, Last I' format."""
    if not s:
        return ""
    s = normalize_text(s)
    # Split on comma (with or without space), rejoin with ', '
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return ", ".join(parts)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    publications = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = normalize_text(row[col["Title"]] or "")
        if not title:
            continue

        # Parse themes from semicolon-separated string
        themes_raw = normalize_text(row[col["Themes"]] or "")
        themes = [t.strip() for t in themes_raw.split(";") if t.strip()]

        # Parse year as integer
        year = row[col["Year"]]
        if year is not None:
            year = int(year)

        # DOI: store as string or null
        doi = row[col["DOI"]]
        if doi:
            doi = str(doi).strip()
        else:
            doi = None

        # Thumbnail and PDF: null if empty
        thumbnail = row[col["Thumbnail"]]
        if thumbnail:
            thumbnail = str(thumbnail).strip()
        else:
            thumbnail = None

        pdf_url = row[col["PdfFile"]]
        if pdf_url:
            pdf_url = str(pdf_url).strip()
        else:
            pdf_url = None

        pub = {
            "title": title,
            "authors": normalize_authors(row[col["Authors"]] or ""),
            "year": year,
            "journal": normalize_text(row[col["Journal"]] or ""),
            "doi": doi,
            "category": (row[col["Category"]] or "").strip().lower(),
            "themes": themes,
            "thumbnail": thumbnail,
            "pdfUrl": pdf_url,
        }
        publications.append(pub)

    # Preserve Excel row order (user-curated ordering)

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(publications, f, indent=2, ensure_ascii=False)

    print(f"Generated {JSON_PATH} with {len(publications)} publications.")


if __name__ == "__main__":
    main()
